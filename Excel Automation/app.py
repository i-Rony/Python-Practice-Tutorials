import openpyxl as xl

workbook = xl.load_workbook("YOUR_FILE.xlsx")
sheet = workbook["Sheet1"]

for row in range(2, sheet.max_row + 1):
    physics = sheet.cell(row, 3).value
    maths = sheet.cell(row, 4).value
    history = sheet.cell(row, 5).value
    geography = sheet.cell(row, 6).value
    biology = sheet.cell(row, 7).value
    chemistry = sheet.cell(row, 8).value

    tot = physics + maths + history + geography + biology + chemistry
    tot_cell = sheet.cell(row, 9)
    tot_cell.value = tot

    per = round(tot/6)
    per_cell = sheet.cell(row, 10)
    per_cell.value = per

workbook.save("YOUR_FINAL_FILENAME.xlsx")